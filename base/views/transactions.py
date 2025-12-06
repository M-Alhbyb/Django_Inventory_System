from django.shortcuts import render, redirect
from django.db.models import Q
from base.models import Transaction, User
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.contrib import messages
from tablib import Dataset
from base.resources import TransactionResource


def transactions_view(request):
    # Get all transactions
    transactions = Transaction.objects.select_related('user').prefetch_related('items__product').order_by('-date')
    
    # Handle search
    q = request.GET.get('q')
    if q:
        transactions = transactions.filter(
            Q(id__icontains=q) |
            Q(user__username__icontains=q) |
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q)
        )
    
    # Handle transaction type filter
    transaction_type = request.GET.get('type')
    if transaction_type and transaction_type != 'all':
        transactions = transactions.filter(type=transaction_type)
    
    # Handle user type filter
    user_type = request.GET.get('user_type')
    if user_type and user_type != 'all':
        transactions = transactions.filter(user__user_type=user_type)
    
    # Get statistics
    total_transactions = transactions.count()
    total_take = transactions.filter(type='take').count()
    total_payment = transactions.filter(type='payment').count()
    total_restore = transactions.filter(type='restore').count()
    total_fees = transactions.filter(type='fees').count()

    # Handle POST requests
    if request.method == 'POST':
        # Check if it's export or import
        if 'export' in request.POST:
            return export_transactions(transactions)
        elif 'import' in request.POST and request.FILES.get('file'):
            return import_transactions(request, transactions)

    paginator = Paginator(transactions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'total_transactions': total_transactions,
        'total_take': total_take,
        'total_payment': total_payment,
        'total_restore': total_restore,
        'total_fees': total_fees,
        'current_type': transaction_type if transaction_type else 'all',
        'current_user_type': user_type if user_type else 'all',
        'search_query': q if q else '',
    }
    
    return render(request, 'transactions.html', context)


def export_transactions(queryset):
    """Export transactions using django-import-export"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    resource = TransactionResource()
    dataset = resource.export(queryset)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="transactions.xlsx"'
    
    # Convert to xlsx with styling
    from openpyxl import Workbook
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "المعاملات"
    ws.sheet_view.rightToLeft = True  # RTL for Arabic
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    row_fill_light = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    row_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Column widths
    column_widths = {'A': 15, 'B': 20, 'C': 18, 'D': 22, 'E': 40, 'F': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Write headers
    headers = dataset.headers if dataset.headers else ['رقم المعاملة', 'المستخدم', 'نوع المعاملة', 'التاريخ', 'العناصر', 'المجموع']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 30
    
    # Write data rows
    for row_num, row in enumerate(dataset, start=2):
        row_fill = row_fill_light if row_num % 2 == 0 else row_fill_white
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.fill = row_fill
            if col_num in [1, 3, 6]:
                cell.alignment = center_alignment
            else:
                cell.alignment = cell_alignment
    
    # Save to response
    wb.save(response)
    return response


def import_transactions(request, queryset):
    """Import transactions using django-import-export"""
    resource = TransactionResource()
    dataset = Dataset()
    
    uploaded_file = request.FILES['file']
    file_format = uploaded_file.name.split('.')[-1].lower()
    
    try:
        if file_format == 'csv':
            imported_data = dataset.load(uploaded_file.read().decode('utf-8'), format='csv')
        elif file_format in ['xlsx', 'xls']:
            imported_data = dataset.load(uploaded_file.read(), format='xlsx')
        else:
            messages.error(request, 'صيغة الملف غير مدعومة. يرجى استخدام ملف CSV أو Excel.')
            return redirect('base:transactions')
        
        # Dry run first to check for errors
        result = resource.import_data(dataset, dry_run=True)
        
        if result.has_errors():
            error_messages = []
            for row_errors in result.row_errors():
                row_num, errors = row_errors
                for error in errors:
                    error_messages.append(f"الصف {row_num}: {error.error}")
            messages.error(request, f'حدثت أخطاء أثناء الاستيراد: {"; ".join(error_messages[:5])}')
        else:
            # Actually import the data
            result = resource.import_data(dataset, dry_run=False)
            messages.success(request, f'تم استيراد {result.totals["new"]} معاملة جديدة بنجاح!')
            
    except Exception as e:
        messages.error(request, f'فشل الاستيراد: {str(e)}')
    
    return redirect('base:transactions')
